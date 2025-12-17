import streamlit as st
import pandas as pd
import requests
from io import BytesIO
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ===============================
# CONFIGURACIÓN GENERAL
# ===============================
st.set_page_config(
    page_title="Cumplimiento – Rechazos BCP",
    layout="wide"
)

API_URL = "https://q6capnpv09.execute-api.us-east-1.amazonaws.com/OPS/kpayout/v1/payout_process/reject_invoices_batch"

HEADERS = {
    # "Authorization": "Bearer TU_TOKEN"
}

CODIGO_RECHAZO = "R016"
DESCRIPCION_RECHAZO = "CUENTA INVÁLIDA"

# ===============================
# RUTA SEGURA PLANTILLA
# ===============================
BASE_DIR = Path(__file__).resolve().parent
PLANTILLA_DD_PATH = BASE_DIR / "plantillas" / "Formato_Due_Diligence_Template.xlsx"

# ===============================
# FUNCIONES
# ===============================
def generar_excel_rechazo(referencias):
    df = pd.DataFrame({
        "Referencia": referencias,
        "Estado": ["Rechazada"] * len(referencias),
        "Codigo de Rechazo": [CODIGO_RECHAZO] * len(referencias),
        "Descripcion de Rechazo": [DESCRIPCION_RECHAZO] * len(referencias)
    })

    buffer = BytesIO()
    df.to_excel(buffer, index=False, engine="openpyxl")
    buffer.seek(0)

    wb = load_workbook(buffer)
    ws = wb.active

    # Forzar referencia como texto
    for col in ws.iter_cols(min_col=1, max_col=1, min_row=2):
        for cell in col:
            cell.number_format = "@"

    # Ajustar ancho columnas
    for column_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 3

    buffer_final = BytesIO()
    wb.save(buffer_final)
    buffer_final.seek(0)

    return buffer_final


def enviar_rechazo_api(buffer_excel):
    files = {
        "edt": (
            "RechazoBCP.xlsx",
            buffer_excel,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    }

    response = requests.post(
        API_URL,
        headers=HEADERS,
        files=files,
        timeout=60
    )

    return response


def generar_formato_due_diligence(df):
    if not PLANTILLA_DD_PATH.exists():
        st.error("❌ No se encontró la plantilla Due Diligence.")
        st.stop()

    fecha_excel = datetime.now().strftime("%d/%m/%Y")
    fecha_archivo = datetime.now().strftime("%d.%m.%y")

    wb = load_workbook(PLANTILLA_DD_PATH)
    ws = wb.active  # o wb["Due Diligence"]

    # Ajusta si tu plantilla usa otra celda
    ws["D10"] = fecha_excel

    fila_inicio = 13

    # Limpiar datos previos
    for row in ws.iter_rows(min_row=fila_inicio, max_col=5):
        for cell in row:
            cell.value = None

    # Insertar clientes seleccionados
    for i, row in enumerate(df.itertuples(), start=0):
        ws[f"C{fila_inicio + i}"] = row.DOCUMENTO
        ws[f"D{fila_inicio + i}"] = str(row.NUMERO_DOCUMENTO)
        ws[f"E{fila_inicio + i}"] = row.NOMBRE

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer, f"Formato Due Dilligence {fecha_archivo}.xlsx"


# ===============================
# INTERFAZ
# ===============================
st.title("🚨 Cumplimiento – Rechazo de Clientes (>30K)")
st.write("Carga archivos Excel, selecciona clientes y ejecuta rechazo vía Postman.")

st.info(f"📄 Plantilla Due Diligence encontrada: {PLANTILLA_DD_PATH.exists()}")

uploaded_files = st.file_uploader(
    "📂 Cargar uno o más archivos Excel",
    type=["xlsx", "xls"],
    accept_multiple_files=True
)

if uploaded_files:
    dataframes = []

    for file in uploaded_files:
        try:
            df = pd.read_excel(file)

            columnas_interes = df.iloc[:, [1, 2, 3, 8, 12]].copy()
            columnas_interes.columns = [
                "DOCUMENTO",
                "NUMERO_DOCUMENTO",
                "NOMBRE",
                "REFERENCIA",
                "MONTO"
            ]

            columnas_interes["MONTO"] = pd.to_numeric(
                columnas_interes["MONTO"], errors="coerce"
            )

            columnas_interes["REFERENCIA"] = columnas_interes["REFERENCIA"].astype(str)

            filtrado = columnas_interes[columnas_interes["MONTO"] > 30000]
            filtrado["Archivo_Origen"] = file.name

            dataframes.append(filtrado)

        except Exception as e:
            st.error(f"Error procesando {file.name}: {e}")

    if dataframes:
        resultado_final = pd.concat(dataframes, ignore_index=True)

        st.subheader("📋 Clientes Observados (>30K)")

        resultado_final["Seleccionar"] = False

        editable_df = st.data_editor(
            resultado_final,
            use_container_width=True,
            num_rows="fixed",
            column_config={
                "Seleccionar": st.column_config.CheckboxColumn(
                    "Rechazar",
                    help="Selecciona clientes a rechazar"
                )
            }
        )

        seleccionados = editable_df[editable_df["Seleccionar"]]

        # ===============================
        # EXCEL DE EVIDENCIAS (TABLA STREAMLIT)
        # ===============================
        if not seleccionados.empty:
            buffer_evidencias = BytesIO()
            seleccionados.drop(columns=["Seleccionar"]).to_excel(
                buffer_evidencias,
                index=False,
                engine="openpyxl"
            )
            buffer_evidencias.seek(0)

            st.download_button(
                "⬇️ Descargar Excel de Evidencias (Seleccionados)",
                data=buffer_evidencias,
                file_name="Evidencias_Clientes_Rechazados.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # ===============================
        # DUE DILIGENCE
        # ===============================
        if not seleccionados.empty:
            excel_dd, nombre_dd = generar_formato_due_diligence(seleccionados)

            st.download_button(
                "📄 Descargar Formato Due Diligence (Seleccionados)",
                data=excel_dd,
                file_name=nombre_dd,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # ===============================
        # RECHAZO VIA POSTMAN / API
        # ===============================
        if not seleccionados.empty:
            st.subheader("🚫 Rechazo vía Postman / API")

            if st.button("Ejecutar Rechazo de Clientes Seleccionados"):
                referencias = seleccionados["REFERENCIA"].tolist()

                excel_api = generar_excel_rechazo(referencias)
                response = enviar_rechazo_api(excel_api)

                if response.status_code in (200, 201):
                    st.success("✅ Rechazo ejecutado correctamente vía Postman.")
                else:
                    st.error("❌ Error en rechazo vía API.")
                    st.write("Status:", response.status_code)
                    st.text(response.text)
        else:
            st.info("ℹ️ Selecciona al menos un cliente para continuar.")
